# services/report_service.py
import csv
from datetime import datetime
from typing import Any

import openpyxl
from django.db.models import Q, Sum
from django.http import HttpRequest, HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1

from apps.inventory.models import InventoryMovement

HEADERS = [
    "Fecha",
    "Tipo Movimiento",
    "Insumo",
    "Lote",
    "Concepto",
    "Entrada",
    "Salida",
    "Balance",
    "Estado",
    "Creado por",
    "Costo Unitario",
    "N° Orden",
]


class ExcelExportService:
    def __init__(self, report_data: dict[str, Any], filename_prefix: str = "report"):
        self.report_data = report_data
        self.filename_prefix = filename_prefix
        self.wb = None
        self.ws = None

    def generate(self) -> HttpResponse:
        self._create_workbook()
        self._add_headers()
        self._add_data()
        self._add_summary()
        self._adjust_column_widths()

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"{self.filename_prefix}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        self.wb.save(response)
        return response

    def _create_workbook(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Inventory Movements"

    def _add_headers(self):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(HEADERS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _add_data(self):
        number_format = FORMAT_NUMBER_COMMA_SEPARATED1

        for row_idx, item in enumerate(self.report_data["data"], 2):
            is_increment = item.get("is_increment", False)
            quantity = item.get("quantity", 0)
            inbound = quantity if is_increment else 0
            outbound = quantity if not is_increment else 0
            balance = item.get("after_stock", 0)

            created_at = item.get("created_at")
            date_str = datetime.fromisoformat(created_at).strftime("%Y-%m-%d")

            data = [
                date_str,
                item.get("movement_type", ""),
                item.get("product_name", ""),
                item.get("batch_number", ""),
                item.get("concept", ""),
                inbound,
                outbound,
                balance,
                item.get("status", ""),
                item.get("created_by", "system"),
                float(item.get("unit_cost", 0)),
                item.get("order_number", ""),
            ]

            for col, value in enumerate(data, 1):
                cell = self.ws.cell(row=row_idx, column=col, value=value)

                if col in [6, 7, 8, 11]:
                    cell.number_format = number_format

                if col == 6 and inbound > 0:
                    cell.font = Font(color="27AE60")
                if col == 7 and outbound > 0:
                    cell.font = Font(color="E74C3C")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            for col in range(1, len(HEADERS) + 1):
                self.ws.cell(row=row_idx, column=col).border = thin_border

    def _add_summary(self):
        if not self.report_data["data"]:
            return

        summary_row = len(self.report_data["data"]) + 3
        summary = self.report_data["summary"]

        self.ws.merge_cells(f"A{summary_row}:E{summary_row}")
        cell = self.ws.cell(row=summary_row, column=1, value="RESUMEN DEL PERÍODO")
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

        metrics = [
            (6, f"Entradas: {summary.get('total_entries', 0)}"),
            (7, f"Salidas: {summary.get('total_exits', 0)}"),
            (8, f"Balance Final: {summary.get('net_balance', 0)}"),
            (10, f"Total Movimientos: {summary.get('total_movements', 0)}"),
        ]

        for col, text in metrics:
            cell = self.ws.cell(row=summary_row, column=col, value=text)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    def _adjust_column_widths(self):
        widths = {
            "A": 20,
            "B": 18,
            "C": 35,
            "D": 15,
            "E": 35,
            "F": 14,
            "G": 14,
            "H": 14,
            "I": 15,
            "J": 18,
            "K": 18,
            "L": 15,
        }
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

        self.ws.row_dimensions[1].height = 25


class CSVExportService:
    def __init__(self, report_data: dict[str, Any], filename_prefix: str = "report"):
        self.report_data = report_data
        self.filename_prefix = filename_prefix

    def generate(self) -> HttpResponse:
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        filename = f"{self.filename_prefix}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response.write("\ufeff")  # BOM para UTF-8

        writer = csv.writer(response, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(HEADERS)

        for item in self.report_data["data"]:
            entrada = item.get("quantity") if item.get("is_increment") else 0
            salida = item.get("quantity") if not item.get("is_increment") else 0
            writer.writerow(
                [
                    item.get("created_at").format("%Y-%m-%d"),
                    item.get("movement_type"),
                    item.get("product_name"),
                    item.get("batch_number"),
                    item.get("concept"),
                    entrada,
                    salida,
                    item.get("after_stock"),
                    item.get("status"),
                    item.get("created_by"),
                    f"{float(item.get('unit_cost', 0)):.2f}",
                    item.get("order_number"),
                ]
            )

        writer.writerow([])
        s = self.report_data["summary"]
        writer.writerow(
            [
                "RESUMEN",
                "",
                "",
                "",
                "",
                f"Entradas: {s['total_entries']}",
                f"Salidas: {s['total_exits']}",
                "",
                f"Balance: {s['net_balance']}",
            ]
        )

        return response


class MovementFilterService:
    @staticmethod
    def extract_filters(request: HttpRequest) -> dict[str, Any]:
        if request.method == "POST":
            return MovementFilterService._extract_from_post(request)
        return MovementFilterService._extract_from_get(request)

    @staticmethod
    def _extract_from_post(request: HttpRequest) -> dict[str, Any]:
        filters = {}
        field_mapping = {
            "status": "status",
            "movement_type": "movement_type",
            "search": "search",
            "created_at_from": "created_at_from",
            "created_at_to": "created_at_to",
            "batch": "batch_id",
            "supply": "supply_id",
        }

        for param, filter_key in field_mapping.items():
            value = request.POST.get(param)
            if value:
                if "date" in filter_key:
                    filters[filter_key] = parse_date(value)
                else:
                    filters[filter_key] = value

        return filters

    @staticmethod
    def _extract_from_get(request: HttpRequest) -> dict[str, Any]:
        filters = {}

        if status := request.GET.get("status"):
            filters["status"] = status
        if movement_type := request.GET.get("movement_type"):
            filters["movement_type"] = movement_type
        if search := request.GET.get("search"):
            filters["search"] = search
        if date_from := request.GET.get("created_at_from"):
            filters["created_at_from"] = parse_date(date_from)
        if date_to := request.GET.get("created_at_to"):
            filters["created_at_to"] = parse_date(date_to)
        if batch_id := request.GET.get("batch"):
            filters["batch_id"] = batch_id
        if supply_id := request.GET.get("supply"):
            filters["supply_id"] = supply_id

        return filters


class InventoryReportService:
    def generate_movement_report(self, filters: dict[str, Any]) -> dict[str, Any]:
        queryset = self._get_base_queryset()
        queryset = self._apply_filters(queryset, filters)

        return {
            "data": self._serialize_movements(queryset),
            "summary": self._calculate_summary(queryset),
            "total_records": queryset.count(),
            "filters_applied": filters,
            "generated_at": timezone.now(),
        }

    def _get_base_queryset(self):
        return InventoryMovement.objects.select_related(
            "batch", "batch__supply", "inventory_order"
        ).order_by("created_at")

    def _apply_filters(self, queryset, filters):
        if status := filters.get("status"):
            queryset = queryset.filter(status=status)
        if movement_type := filters.get("movement_type"):
            queryset = queryset.filter(movement_type=movement_type)
        if search := filters.get("search"):
            queryset = queryset.filter(
                Q(concept__icontains=search)
                | Q(batch__batch_number__icontains=search)
                | Q(batch__supply__name__icontains=search)
            )
        if date_from := filters.get("created_at_from"):
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to := filters.get("created_at_to"):
            queryset = queryset.filter(created_at__date__lte=date_to)
        if batch_id := filters.get("batch_id"):
            queryset = queryset.filter(batch_id=batch_id)
        if supply_id := filters.get("supply_id"):
            queryset = queryset.filter(batch__supply_id=supply_id)
        return queryset

    def _serialize_movements(self, queryset):
        return [self._serialize_movement(m) for m in queryset]

    def _serialize_movement(self, movement):
        return {
            "id": movement.id,
            "external_id": str(movement.external_id),
            "created_at": movement.created_at.strftime("%Y-%m-%d"),
            "movement_type": movement.get_movement_type_display(),
            "concept": movement.concept,
            "quantity": movement.quantity,
            "is_increment": movement.is_increment,
            "previous_stock": movement.previous_stock,
            "after_stock": movement.after_stock,
            "status": movement.get_status_display(),
            "created_by": movement.created_by,
            "unit_cost": float(movement.unit_cost_at_movement)
            if movement.unit_cost_at_movement
            else 0,
            "batch_number": movement.batch.batch_number if movement.batch else None,
            "product_name": movement.batch.supply.name
            if movement.batch and movement.batch.supply
            else None,
            "order_number": movement.inventory_order.order_number
            if movement.inventory_order
            else None,
        }

    def _calculate_summary(self, queryset):
        total_entries = (
            queryset.filter(is_increment=True).aggregate(total=Sum("quantity"))["total"] or 0
        )

        total_exits = (
            queryset.filter(is_increment=False).aggregate(total=Sum("quantity"))["total"] or 0
        )

        return {
            "total_entries": total_entries,
            "total_exits": total_exits,
            "net_balance": total_entries - total_exits,
            "total_movements": queryset.count(),
        }
